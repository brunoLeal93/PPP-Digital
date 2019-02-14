from openpyxl import load_workbook
import time
#from mongoengine import connect
from pymongo import MongoClient
#import sqlite3

client = MongoClient("mongodb://admin:admin@cluster0-shard-00-00-umdst.mongodb.net:27017,cluster0-shard-00-01-umdst.mongodb.net:27017,cluster0-shard-00-02-umdst.mongodb.net:27017/test?ssl=true&replicaSet=Cluster0-shard-0&authSource=admin&retryWrites=true")
db = client.BasePPPDigital

def importaBase():

    coll = db.Cotacao

    print("Abrindo Arquivo")
    wb = load_workbook('MODELBASE.xlsx', read_only=True)
    sheet = wb['Plan2']
    print("Planilha aberta")
    row= 2
    vet = []
    i= 1

    #print(sheet.cell(column=1, row=row).value)

    while row < 200000:
        print("row: {}".format(row))
        rt = []
        delay = []
        aux = row

<<<<<<< HEAD
        if sheet.cell(column=2, row=aux).value == sheet.cell(column=2, row=aux+1).value:

            while sheet.cell(column=2, row=aux).value == sheet.cell(column=2, row=aux+1).value:

                if sheet.cell(column=12, row=aux).value == 'RT':

                    doc = {
                        'cod_pag': sheet.cell(column=13, row=aux).value,
                        'cod_serv': sheet.cell(column=14, row=aux).value,
                        'desc_serv': sheet.cell(column=15, row=aux).value,
                        'vlr_serv': sheet.cell(column=16, row=aux).value,
                        'vlr_fee': sheet.cell(column=17, row=aux).value
                    }
                    rt.append(doc)
                else:
                    doc = {
                        'cod_pag': sheet.cell(column=13, row=aux).value,
                        'cod_serv': sheet.cell(column=14, row=aux).value,
                        'desc_serv': sheet.cell(column=15, row=aux).value,
                        'vlr_serv': sheet.cell(column=16, row=aux).value,
                        'vlr_fee': sheet.cell(column=17, row=aux).value
                    }
                    delay.append(doc)

                aux = aux + 1

            doc= {
                    'tipo': sheet.cell(column=1, row=aux).value,
                    'cod_ativo':sheet.cell(column=2, row=aux).value,
                    'cod_ativo_bolsa':sheet.cell(column=3, row=aux).value,
                    'desc_ativo':sheet.cell(column=4, row=aux).value,
                    'contrato_ativo':sheet.cell(column=5, row=aux).value,
                    'extensao_ativo':sheet.cell(column=6, row=aux).value,
                    'fonte_ativo':sheet.cell(column=7, row=aux).value.upper(),
                    'seg_fonte_ativo':sheet.cell(column=8, row=aux).value,
                    'merc_ativo':sheet.cell(column=9, row=aux).value,
                    'classe_ativo':sheet.cell(column=10, row=aux).value,
                    'inf_disp':sheet.cell(column=11, row=aux).value,
                    'servicos': {
                        'rt': rt,
                        'delay': delay,
                    }
            }

            vet.append(doc)
        else:
            if sheet.cell(column=12, row=row).value == 'RT':

                doc = {
                    'cod_pag': sheet.cell(column=13, row=row).value,
                    'cod_serv': sheet.cell(column=14, row=row).value,
                    'desc_serv': sheet.cell(column=15, row=row).value,
                    'vlr_serv': sheet.cell(column=16, row=row).value,
                    'vlr_fee': sheet.cell(column=17, row=row).value
=======
            if sheet.cell(column=12, row=aux).value == 'RT':

                doc = {
                    'cod_pag': sheet.cell(column=13, row=aux).value,
                    'cod_serv': sheet.cell(column=14, row=aux).value,
                    'desc_serv': sheet.cell(column=15, row=aux).value,
                    'vlr_serv': sheet.cell(column=16, row=aux).value,
                    'vlr_fee': sheet.cell(column=17, row=aux).value
>>>>>>> 17c52eae9d38ed005fde80c78ac4d101e17c5630
                }
                rt.append(doc)
            else:
                doc = {
<<<<<<< HEAD
                    'cod_pag': sheet.cell(column=13, row=row).value,
                    'cod_serv': sheet.cell(column=14, row=row).value,
                    'desc_serv': sheet.cell(column=15, row=row).value,
                    'vlr_serv': sheet.cell(column=16, row=row).value,
                    'vlr_fee': sheet.cell(column=17, row=row).value
                }
                delay.append(doc)

            doc = {
                'tipo': sheet.cell(column=1, row=row).value,
                'cod_ativo': sheet.cell(column=2, row=row).value,
                'cod_ativo_bolsa': sheet.cell(column=3, row=row).value,
                'desc_ativo': sheet.cell(column=4, row=row).value,
                'contrato_ativo': sheet.cell(column=5, row=row).value,
                'extensao_ativo': sheet.cell(column=6, row=row).value,
                'fonte_ativo': sheet.cell(column=7, row=row).value.upper(),
                'seg_fonte_ativo': sheet.cell(column=8, row=row).value,
                'merc_ativo': sheet.cell(column=9, row=row).value,
                'classe_ativo': sheet.cell(column=10, row=row).value,
                'inf_disp': sheet.cell(column=11, row=row).value,
=======
                    'cod_pag': sheet.cell(column=13, row=aux).value,
                    'cod_serv': sheet.cell(column=14, row=aux).value,
                    'desc_serv': sheet.cell(column=15, row=aux).value,
                    'vlr_serv': sheet.cell(column=16, row=aux).value,
                    'vlr_fee': sheet.cell(column=17, row=aux).value
                }
                delay.append(doc)

            aux = aux + 1

        doc= {
                'tipo': sheet.cell(column=1, row=aux).value,
                'cod_ativo':sheet.cell(column=2, row=aux).value,
                'cod_ativo_bolsa':sheet.cell(column=3, row=aux).value,
                'desc_ativo':sheet.cell(column=4, row=aux).value,
                'contrato_ativo':sheet.cell(column=5, row=aux).value,
                'extensao_ativo':sheet.cell(column=6, row=aux).value,
                'fonte_ativo':sheet.cell(column=7, row=aux).value.upper(),
                'seg_fonte_ativo':sheet.cell(column=8, row=aux).value,
                'merc_ativo':sheet.cell(column=9, row=aux).value,
                'classe_ativo':sheet.cell(column=10, row=aux).value,
                'inf_disp':sheet.cell(column=11, row=aux).value,
>>>>>>> 17c52eae9d38ed005fde80c78ac4d101e17c5630
                'servicos': {
                    'rt': rt,
                    'delay': delay,
                }
            }
            vet.append(doc)
        resto = row//500

        if row % 300 == 0:
            coll.insert_many(vet)
            vet=[]
            print("Importacação #{}".format(i))
            i+=1
        row = row +1

        if sheet.cell(column=1, row=row).value == None:
            break

    print("Importação Concluída")

def cotVisuPrin():
    coll = db.Cotacao

    vet = []

    for doc in coll.find({}, {"_id": 0, "desc_pag": 0, "desc_ticker": 0, "desc_pag": 0, "ticker": 0}):

        vet.append(doc)

    return vet


ini = time.time()
importaBase()
fim = time.time()
print(fim - ini)
#consultaBase()

#a = cotVisuPrin()
#print(a)
